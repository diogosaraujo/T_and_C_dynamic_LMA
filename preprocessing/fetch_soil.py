#!/usr/bin/env python3
"""Build depth-resolved soil profiles and column depths for the AmeriFlux stations.

Two things T&C needs and cannot derive: the texture profile (Psan/Pcla/Porg per
layer, which Soil_parameters turns into Osat/Ks/alpVG/nVG via Saxton & Rawls) and
the depth of the soil column itself. US_xRM's 1 m free-draining column is a
placeholder; inheriting it for a deep-rooted forest biases the site toward drought
stress and under-ET/GPP, which would contaminate the very LMA signal this project
is measuring.

SOURCE ORDER, first hit wins, recorded per field:

    1. AmeriFlux BADM   in-situ, where the site reported SOIL_TEX / SOIL_CHEM
    2. SSURGO via SDA   USDA field survey. The only source giving a field-observed
                        DEPTH TO BEDROCK rather than a modelled estimate, from the
                        same profile the texture comes from.

    3. POLARIS          30 m CONUS, itself disaggregated from SSURGO. Gap filler.
    4. SoilGrids        250 m global. Last resort, and only properties -- SoilGrids
                        2.0 dropped the bedrock layers, so it cannot supply depth.

COLUMN DEPTH is depth to bedrock; rooting depth is a separate input (Schenk &
Jackson), and the two meet only in the constraint ZR95 <= column depth, applied
when the .mat files are built. corestrictions catalogues layers restricting roots
AND/OR water, so only its BEDROCK kinds end the column. Densic material, fragipans
and abrupt textural changes lie within the soil: the survey keeps describing
horizons 80-140 cm below them (US-Ho1 70 vs 165 cm, US-MOz 28 vs 152), whereas at a
bedrock contact it stops dead (US-MtB 33/33, US-CPk 56/56). They are recorded, never
used as a floor.

THE FOREST FLOOR is not part of the column. SSURGO names every horizon, so O
horizons (Oi/Oe/Oa litter and duff, desgnmaster 'O') are identified from the survey
rather than guessed at from organic content, dropped, and the remaining depths
re-zeroed on the mineral surface -- including the bedrock contact.

This is not merely a Saxton & Rawls workaround. T&C models the forest floor
ALREADY, and prognostically: litter biomass BLit from the carbon cycle gives
Llitter = Sllit*BLit, hence a cover fraction CLitter, an interception store
In_Litter with its own evaporation ELitter in the latent heat flux, and r_litter,
which is added to the soil-evaporation resistance in Heat_fluxes.m as a mulch
effect. Feeding SSURGO's O horizon into Porg would double-count the forest floor:
once as a soil layer with impossible hydraulics, and again as the litter layer the
model grows itself. Porg describes MINERAL soil, and reaches only Soil_parameters.
The papers do the same thing, using Porg = 0.01 taken
from 8 g/kg of organic carbon measured in the 0-6 cm MINERAL horizon at Lucky
Hills; MOD_PARAM_US_xRM.m's Porg = 0.07 is likewise an A-horizon value.

ORGANIC MATTER is the field most easily got wrong, because every source states it
differently: SSURGO om_r is organic matter as a percent, POLARIS om is log10 of
percent, SoilGrids soc is SOC in dg/kg and needs the x1.72 conversion to organic
matter. All three are normalised to Porg as a FRACTION here, which is what
Soil_parameters expects (it multiplies by 100 internally).

LAYERING. Texture is mapped onto the site's Zs mesh and Soil_parameters is run per
layer, rather than one triple replicated with *ones(1,ms). No source resolves finer
than 5 cm, so the top three model layers (0-1, 1-2, 2-5 cm) necessarily share a
value -- the profile is piecewise constant with a handful of distinct values, not
13 independent ones. Below the described depth, the deepest observed layer is
carried down, and every such layer is flagged `extrapolated`.

    python fetch_soil.py --probe          # coverage only: what SDA returns, no files
    python fetch_soil.py                  # build the profiles
    python fetch_soil.py --report         # summarise what was obtained for 101 sites
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import os
import sys
import time
import urllib.error
import urllib.request
from collections import Counter, defaultdict
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
PREPROC = Path(__file__).resolve().parent
DEFAULT_SITE_LISTS = [
    REPO_ROOT / "T&C" / "dynamic_lma_test" / "deciduous_ameriflux.csv",
    REPO_ROOT / "T&C" / "dynamic_lma_test" / "evergreen_ameriflux.csv",
]
DEFAULT_EXCLUDED = PREPROC / "excluded_stations.csv"
INPUT_ROOT = Path(os.environ.get("TC_INPUT_DATA",
                                 "/vol_efthymios/NFS07/dd1136/T_and_C/input_data"))
DEFAULT_OUT = INPUT_ROOT / "soil"
DEFAULT_ROOT_DEPTH = INPUT_ROOT / "root_depth" / "root_depth_schenk_jackson.csv"
DEFAULT_BADM_DIR = INPUT_ROOT / "ameriflux"

SDA_URL = "https://sdmdataaccess.sc.egov.usda.gov/Tabular/post.rest"
POLARIS_BASE = "http://hydrology.cee.duke.edu/POLARIS/PROPERTIES/v1.0"
SOILGRIDS_URL = "https://rest.isric.org/soilgrids/v2.0/properties/query"

# The six standard intervals POLARIS and SoilGrids both publish, in cm.
STD_INTERVALS = [(0, 5), (5, 15), (15, 30), (30, 60), (60, 100), (100, 200)]

# US_xRM's graded mesh, in mm. Kept as the top of every site's column so the
# near-surface resolution that drives evaporation is not lost; deeper nodes are
# appended per site to reach the column depth.
BASE_ZS = [0, 10, 20, 50, 100, 150, 200, 300, 400, 500, 600, 700, 800, 1000]

MISSING = {"", "-9999", "-9999.0", "NA", "N/A", "NaN", "nan", None}
# The reskind values that ARE bedrock, and so mark the bottom of the soil column.
# Everything else corestrictions reports -- densic material, fragipans, abrupt
# textural changes -- lies within the soil and does not end it. Counts across the
# 101 stations (job 35576): Lithic 8, Paralithic 10 | Densic 15, Fragipan 3,
# Abrupt textural change 1 | no restriction 64.
BEDROCK_KINDS = {"lithic bedrock", "paralithic bedrock", "densic bedrock"}

# Porg used where a source reports no organic matter for a horizon: 0.1%. See the
# note in process() -- filling from the topsoil value instead would be badly wrong.
ORG_FLOOR = 0.001


# --------------------------------------------------------------------------- io

def read_excluded(path: Path | None) -> dict[str, str]:
    out: dict[str, str] = {}
    if not path or not Path(path).is_file():
        return out
    with open(path, newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            sid = (row.get("station_id") or "").strip()
            if sid:
                out[sid] = (row.get("reason") or "excluded").strip()
    return out


def read_stations(paths, wanted: set[str] | None) -> list[dict]:
    stations, seen = [], set()
    for path in paths:
        if not Path(path).is_file():
            print(f"  ! site list not found: {path}", file=sys.stderr)
            continue
        with open(path, newline="", encoding="utf-8-sig") as fh:
            for row in csv.DictReader(fh):
                sid = (row.get("StationID") or "").strip()
                if not sid or sid in seen or (wanted and sid not in wanted):
                    continue
                try:
                    lat, lon = float(row["Lat"]), float(row["Lon"])
                except (KeyError, TypeError, ValueError):
                    print(f"  ! {sid}: unusable Lat/Lon, skipped", file=sys.stderr)
                    continue
                seen.add(sid)
                stations.append({"station_id": sid, "lat": lat, "lon": lon,
                                 "forest_type": (row.get("ForestType") or "").strip().lower()})
    return sorted(stations, key=lambda s: s["station_id"])


def read_root_depth(path: Path) -> dict[str, float]:
    """station -> ZR95 in mm, so the column can be kept at least as deep."""
    out: dict[str, float] = {}
    if not Path(path).is_file():
        print(f"  ! root depth table not found: {path} -- ZR95 will not be enforced",
              file=sys.stderr)
        return out
    with open(path, newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            sid = (row.get("station_id") or row.get("StationID") or "").strip()
            for key in ("ZR95_H_mm", "ZR95_mm", "zr95_mm", "ZR95", "root_depth_mm"):
                if sid and row.get(key) not in MISSING:
                    try:
                        out[sid] = float(row[key])
                    except (TypeError, ValueError):
                        continue
                    break
    return out


# ------------------------------------------------------------------- 1. BADM

def read_badm_rows(path: Path) -> list[dict]:
    if path.suffix.lower() in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            return []
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(rows, [])]
        out = [dict(zip(header, [("" if c is None else str(c).strip()) for c in r]))
               for r in rows]
        wb.close()
        return out
    with open(path, newline="", encoding="utf-8-sig") as fh:
        return list(csv.DictReader(fh))


def badm_profile(badm_dir: Path, station: str) -> tuple[list[dict], str]:
    """In-situ horizons from a station's BIF file, if it reported soil texture.

    BIF is long-format (SITE_ID, VARIABLE_GROUP, VARIABLE, DATAVALUE), with the
    fields of one observation tied together by GROUP_ID. Sand/silt/clay and the
    depth they refer to therefore have to be regrouped before they mean anything.
    """
    sdir = badm_dir / station
    if not sdir.is_dir():
        return [], ""
    files = [p for p in sorted(sdir.iterdir())
             if p.suffix.lower() in (".xlsx", ".xls", ".csv") and "BIF" in p.name.upper()]
    groups: dict[str, dict[str, str]] = defaultdict(dict)
    used = ""
    for path in files:
        for row in read_badm_rows(path):
            var = (row.get("VARIABLE") or "").strip().upper()
            val = (row.get("DATAVALUE") or "").strip()
            if not var.startswith("SOIL_TEX") or val in MISSING:
                continue
            gid = (row.get("GROUP_ID") or row.get("VARIABLE_GROUP") or var).strip()
            groups[gid][var] = val
            used = path.name

    horizons = []
    for fields in groups.values():
        def num(*names):
            for n in names:
                if n in fields:
                    try:
                        return float(fields[n])
                    except ValueError:
                        pass
            return None
        sand, clay = num("SOIL_TEX_SAND"), num("SOIL_TEX_CLAY")
        if sand is None or clay is None:
            continue
        top = num("SOIL_TEX_DEPTH_TOP", "SOIL_TEX_UPPER_DEPTH") or 0.0
        bot = num("SOIL_TEX_DEPTH_BOTTOM", "SOIL_TEX_LOWER_DEPTH", "SOIL_TEX_DEPTH")
        if bot is None or bot <= top:
            continue
        horizons.append({"top_cm": top, "bot_cm": bot,
                         "sand": sand / 100.0, "clay": clay / 100.0,
                         "org": None, "source": "ameriflux_badm"})
    horizons.sort(key=lambda h: h["top_cm"])
    return horizons, used


# ------------------------------------------------------------------ 2. SSURGO

def sda_query(sql: str, retries: int = 3, timeout: int = 60,
              optional: bool = False) -> list[dict]:
    """POST one SQL statement to Soil Data Access. No authentication needed.

    A 4xx means SDA rejected the SQL -- an unknown column, usually -- and retrying
    cannot help, so it fails immediately and says so rather than reporting 'failed
    after 3 attempts', which reads like a network fault. optional=True downgrades
    any failure to an empty result, for fields that are nice to have: job 35578 lost
    all 101 stations because one cross-check column did not exist.
    """
    body = json.dumps({"query": sql, "format": "JSON+COLUMNNAME"}).encode("utf-8")
    req = urllib.request.Request(SDA_URL, data=body,
                                 headers={"Content-Type": "application/json"})
    last = None
    for attempt in range(retries):
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                payload = json.loads(resp.read().decode("utf-8"))
            table = payload.get("Table") or []
            if not table:
                return []
            head, *rows = table
            return [dict(zip(head, r)) for r in rows]
        except urllib.error.HTTPError as exc:
            if 400 <= exc.code < 500:
                if optional:
                    return []
                raise RuntimeError(
                    f"SDA rejected the query (HTTP {exc.code}) -- check the column "
                    f"names against the SSURGO schema: {sql[:160]}") from exc
            last = exc
            time.sleep(2 * (attempt + 1))
        except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as exc:
            last = exc
            time.sleep(2 * (attempt + 1))
    if optional:
        return []
    raise RuntimeError(f"SDA query failed after {retries} attempts: {last}")


def ssurgo_profile(lat: float, lon: float) -> dict:
    """Horizons, restriction and provenance for the dominant component at a point."""
    mu = sda_query(
        "SELECT mukey FROM SDA_Get_Mukey_from_intersection_with_WktWgs84"
        f"('point({lon:.6f} {lat:.6f})')")
    if not mu or not mu[0].get("mukey"):
        return {"horizons": [], "note": "no SSURGO map unit at this point"}
    mukey = str(mu[0]["mukey"]).strip()

    rows = sda_query(f"""
        SELECT c.cokey, c.compname, c.comppct_r, c.majcompflag,
               ch.hzname, ch.desgnmaster,
               ch.hzdept_r, ch.hzdepb_r, ch.sandtotal_r, ch.silttotal_r,
               ch.claytotal_r, ch.om_r, ch.ksat_r, ch.dbthirdbar_r,
               ch.wthirdbar_r, ch.wfifteenbar_r
        FROM component c
        LEFT OUTER JOIN chorizon ch ON c.cokey = ch.cokey
        WHERE c.mukey = '{mukey}'
        ORDER BY c.comppct_r DESC, ch.hzdept_r ASC""")
    if not rows:
        return {"mukey": mukey, "horizons": [], "note": "map unit has no components"}

    # Dominant component by area percentage. Aggregating across components would
    # average soils that are genuinely different bodies, so the rule is to pick one
    # and record which, along with how much of the map unit it represents.
    def pct(r):
        try:
            return float(r.get("comppct_r") or 0)
        except (TypeError, ValueError):
            return 0.0

    def has_horizons(r):
        return r.get("hzdept_r") not in MISSING and r.get("hzdepb_r") not in MISSING

    # The largest component is often a miscellaneous area -- rock outcrop, water,
    # urban land -- carrying no horizons at all; the probe found six such stations,
    # Hubbard Brook among them. Prefer the largest component that IS described,
    # rather than discarding the map unit and dropping to POLARIS.
    described = [r for r in rows if has_horizons(r)]
    demoted = ""
    if described:
        best = max(described, key=pct)
        top = max(rows, key=pct)
        if str(top["cokey"]).strip() != str(best["cokey"]).strip():
            demoted = (f"largest component '{(top.get('compname') or '?').strip()}' "
                       f"({pct(top):g}%) has no horizons; used "
                       f"'{(best.get('compname') or '?').strip()}' ({pct(best):g}%)")
    else:
        best = max(rows, key=pct)
    cokey = str(best["cokey"]).strip()
    comp = [r for r in rows if str(r["cokey"]).strip() == cokey]

    # The survey names every horizon, so the forest floor never has to be guessed at
    # from its organic content: desgnmaster is the master horizon letter, and O is
    # organic (Oi slightly decomposed litter, Oe moderately, Oa well). L is limnic
    # organic material, R is bedrock -- neither is soil either. Everything reported
    # above 45% organic matter at our stations is an O horizon, so no threshold is
    # needed to identify them; --max-porg remains only as a Saxton & Rawls validity
    # guard on genuine mineral horizons (US-LPH's A horizon really is 10.8%).
    horizons, organic = [], []
    for r in comp:
        try:
            top, bot = float(r["hzdept_r"]), float(r["hzdepb_r"])
        except (TypeError, ValueError, KeyError):
            continue
        if bot <= top:
            continue
        master = (r.get("desgnmaster") or "").strip().upper()
        hzname = (r.get("hzname") or "").strip()
        if master in ("O", "L") or (not master and hzname[:1].upper() == "O"):
            organic.append({"top_cm": top, "bot_cm": bot, "hzname": hzname})
            continue
        if master == "R":                      # bedrock, not soil
            continue
        try:
            sand, clay = float(r["sandtotal_r"]), float(r["claytotal_r"])
        except (TypeError, ValueError, KeyError):
            continue
        om = r.get("om_r")
        try:
            org = float(om) / 100.0            # om_r is organic matter PERCENT
        except (TypeError, ValueError):
            org = None
        horizons.append({"top_cm": top, "bot_cm": bot, "hzname": hzname,
                         "sand": sand / 100.0, "clay": clay / 100.0, "org": org,
                         "ksat_um_s": _f(r.get("ksat_r")),
                         "wthirdbar": _f(r.get("wthirdbar_r")),
                         "wfifteenbar": _f(r.get("wfifteenbar_r")),
                         "source": "ssurgo_sda"})
    horizons.sort(key=lambda h: h["top_cm"])
    organic.sort(key=lambda h: h["top_cm"])

    # Re-zero on the MINERAL surface. SSURGO measures depth from the top of the O
    # horizon; T&C's column is mineral soil, and the papers parameterise it from the
    # mineral surface horizon (Fatichi et al. use Porg = 0.01 from 8 g/kg organic
    # carbon measured at 0-6 cm). Leaving the depths as they are would model the
    # litter thickness as extra mineral soil, so the offset is subtracted from the
    # horizons and, below, from the bedrock contact as well.
    o_thick = 0.0
    if organic and horizons and organic[0]["top_cm"] <= 0.01:
        o_thick = max(o["bot_cm"] for o in organic if o["top_cm"] < horizons[0]["top_cm"] + 0.01)
        o_thick = min(o_thick, horizons[0]["top_cm"])
        for h in horizons:
            h["top_cm"] = max(0.0, h["top_cm"] - o_thick)
            h["bot_cm"] -= o_thick

    # corestrictions is an inventory of layers that restrict roots and/or water. Only
    # the BEDROCK kinds mark the bottom of the soil column, which is the quantity
    # being fetched here -- rooting depth comes from Schenk & Jackson separately. A
    # densic layer, fragipan or abrupt textural change sits WITHIN the soil: the
    # survey keeps describing horizons for another 80-140 cm below them (US-Ho1 70 vs
    # 165 cm, US-PF* 84 vs 203, US-MOz 28 vs 152), whereas at a bedrock contact the
    # description stops dead (US-MtB 33/33, US-CPk 56/56, US-GBT 74/74). Both are
    # returned; only bedrock_cm sets the depth.
    res = sda_query(f"""
        SELECT reskind, resdept_r FROM corestrictions
        WHERE cokey = '{cokey}' ORDER BY resdept_r ASC""")
    restriction, res_depth = "", None
    bedrock, bedrock_depth = "", None
    for r in res:
        kind = (r.get("reskind") or "").strip()
        d = _f(r.get("resdept_r"))
        if d is None or d <= 0:
            continue
        if not restriction:
            restriction, res_depth = kind, d
        if kind.lower() in BEDROCK_KINDS and bedrock_depth is None:
            bedrock, bedrock_depth = kind, max(0.0, d - o_thick)
    # muaggatt.brockdepmin is the MINIMUM bedrock depth across every component of the
    # map unit, while the horizons above come from the dominant one -- so it can
    # describe a different, shallower soil entirely. The probe (job 35515) found it
    # contradicted the dominant component in all 7 cases it fired: US-xRM would have
    # been given a 39 cm column against a profile described to 158 cm, US-CZ4 a 0 cm
    # one. It is therefore recorded for information and never used to set the depth;
    # the dominant component having no corestrictions entry correctly means no
    # restriction within its described profile.
    brock = aws = None
    agg = sda_query(f"SELECT brockdepmin FROM muaggatt WHERE mukey = '{mukey}'")
    if agg:
        d = _f(agg[0].get("brockdepmin"))
        if d is not None and 0 < d < 201:      # 201 is muaggatt's 'deeper than described'
            brock = d
    # Root zone available water storage, mm -- an independent check on whether the
    # column we build holds a plausible amount of water. Queried SEPARATELY and
    # optionally: folding it into the brockdepmin query cost every station in job
    # 35578 when SDA rejected the column name. Nothing depends on it.
    # (rootznemc is deliberately not pulled: a capped commodity-crop rooting depth.)
    extra = sda_query(f"SELECT rootznaws FROM muaggatt WHERE mukey = '{mukey}'",
                      retries=1, optional=True)
    if extra:
        aws = _f(extra[0].get("rootznaws"))

    return {"mukey": mukey, "cokey": cokey,
            "compname": (best.get("compname") or "").strip(),
            "comppct": pct(best), "majcomp": (best.get("majcompflag") or "").strip(),
            "n_components": len({str(r["cokey"]).strip() for r in rows}),
            "restriction": restriction, "restriction_cm": res_depth,
            "bedrock": bedrock, "bedrock_cm": bedrock_depth,
            "o_horizon_cm": o_thick,
            "o_horizons": " ".join(f"{o['hzname']}({o['top_cm']:g}-{o['bot_cm']:g})"
                                   for o in organic),
            "brockdepmin_cm": brock, "rootznaws_mm": aws,
            "horizons": horizons, "note": demoted}


def _f(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# ----------------------------------------------------------------- 3. POLARIS

def polaris_profile(lat: float, lon: float, timeout: int = 60) -> list[dict]:
    """The six standard intervals from the 30 m CONUS product, read over /vsicurl/.

    POLARIS publishes 1x1 degree tiles per variable, statistic and depth. `om` is
    log10 of organic-matter percent, so it is exponentiated before use -- reading it
    as a percent would understate organic matter by orders of magnitude.
    """
    try:
        import rasterio
    except ImportError:
        return []
    os.environ.setdefault("GDAL_DISABLE_READDIR_ON_OPEN", "EMPTY_DIR")
    os.environ.setdefault("CPL_VSIL_CURL_ALLOWED_EXTENSIONS", ".tif")
    lat_lo, lon_lo = math.floor(lat), math.floor(lon)
    tile = f"lat{lat_lo}{lat_lo + 1}_lon{lon_lo}{lon_lo + 1}.tif"

    out = []
    for top, bot in STD_INTERVALS:
        vals = {}
        for var in ("sand", "clay", "om"):
            url = f"/vsicurl/{POLARIS_BASE}/{var}/mean/{top}_{bot}/{tile}"
            try:
                with rasterio.open(url) as src:
                    v = next(src.sample([(lon, lat)]))[0]
                    nod = src.nodata
                if v is None or (nod is not None and v == nod) or not math.isfinite(float(v)):
                    continue
                vals[var] = float(v)
            except Exception:                                      # noqa: BLE001
                continue
        if "sand" not in vals or "clay" not in vals:
            continue
        org = None
        if "om" in vals:
            org = (10.0 ** vals["om"]) / 100.0          # log10(%) -> percent -> fraction
        out.append({"top_cm": top, "bot_cm": bot,
                    "sand": vals["sand"] / 100.0, "clay": vals["clay"] / 100.0,
                    "org": org, "source": "polaris"})
    return out


# --------------------------------------------------------------- 4. SoilGrids

def soilgrids_profile(lat: float, lon: float, timeout: int = 60) -> list[dict]:
    """Global 250 m fallback. Properties only -- SoilGrids 2.0 has no bedrock layer."""
    depths = ["0-5cm", "5-15cm", "15-30cm", "30-60cm", "60-100cm", "100-200cm"]
    url = (f"{SOILGRIDS_URL}?lon={lon}&lat={lat}"
           + "".join(f"&depth={d}" for d in depths)
           + "&property=sand&property=clay&property=soc&value=mean")
    try:
        with urllib.request.urlopen(url, timeout=timeout) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
    except Exception:                                              # noqa: BLE001
        return []

    by_depth: dict[str, dict[str, float]] = defaultdict(dict)
    for layer in (payload.get("properties") or {}).get("layers") or []:
        name = layer.get("name")
        factor = ((layer.get("unit_measure") or {}).get("d_factor")) or 1
        for d in layer.get("depths") or []:
            v = (d.get("values") or {}).get("mean")
            if v is None:
                continue
            by_depth[d.get("label")][name] = float(v) / float(factor)

    out = []
    for label, (top, bot) in zip(depths, STD_INTERVALS):
        vals = by_depth.get(label, {})
        if "sand" not in vals or "clay" not in vals:
            continue
        org = None
        if "soc" in vals:
            # soc arrives in g/kg after the d_factor; organic matter = SOC * 1.72,
            # then g/kg -> percent -> fraction.
            org = (vals["soc"] * 1.72 / 10.0) / 100.0
        out.append({"top_cm": top, "bot_cm": bot,
                    "sand": vals["sand"] / 100.0, "clay": vals["clay"] / 100.0,
                    "org": org, "source": "soilgrids"})
    return out


# ------------------------------------------------------------------ profile ops

def build_zs(depth_mm: float) -> list[float]:
    """Site mesh: the graded US_xRM top, extended to the column depth."""
    zs = [z for z in BASE_ZS if z < depth_mm]
    last = zs[-1] if zs else 0.0
    while last < depth_mm:
        step = 250.0 if last < 2000 else 500.0
        last = min(depth_mm, last + step)
        zs.append(last)
    if zs[-1] != depth_mm:
        zs.append(depth_mm)
    return zs


def layerise(horizons: list[dict], zs: list[float], default_org: float,
             max_porg: float = 0.08) -> tuple[list[dict], float]:
    """Map horizons onto the mesh by layer midpoint, carrying the deepest downward."""
    if not horizons:
        return []
    deepest = horizons[-1]
    described_mm = deepest["bot_cm"] * 10.0
    out = []
    for k in range(len(zs) - 1):
        top, bot = zs[k], zs[k + 1]
        mid_cm = 0.5 * (top + bot) / 10.0
        hit = next((h for h in horizons if h["top_cm"] <= mid_cm < h["bot_cm"]), None)
        extrapolated = hit is None
        if hit is None:
            # Above the shallowest horizon takes the shallowest, below the deepest
            # takes the deepest. Falling through to `deepest` in both directions
            # would give the 0-5 cm layers a C-horizon texture.
            hit = horizons[0] if mid_cm < horizons[0]["top_cm"] else deepest
        org = hit.get("org")
        if org is None:
            org = default_org
        sand, clay = hit["sand"], hit["clay"]

        notes = []
        # Saxton & Rawls (2006) fitted their pedotransfer functions on MINERAL soils
        # with organic matter up to ~8%. SSURGO describes forest-floor O horizons at
        # 45-85% (12 of 95 stations, 8 of them at 0.75-0.85), and pushing those
        # through the regressions produces meaningless porosity and retention -- the
        # equations are simply outside their domain. Cap at the validity limit and
        # record the original. US_xRM's own hand-set Porg = 0.07 is effectively this
        # same decision, made by hand.
        org_raw = org
        if org > max_porg:
            org = max_porg
            notes.append(f"Porg {org_raw:.3f} capped at {max_porg:g} "
                         f"(Saxton & Rawls validity limit)")
        # Soil_parameters computes Psil = 1 - Psan - Pcla - Porg and bails out if it
        # goes negative. Scale the mineral fractions rather than let MATLAB print and
        # return mid-run with undefined outputs.
        if sand + clay + org > 1.0:
            scale = (1.0 - org) / (sand + clay) if (sand + clay) > 0 else 0.0
            sand, clay = sand * scale, clay * scale
            notes.append(f"mineral fractions scaled by {scale:.3f} to keep Psil >= 0")
        adjusted = "; ".join(notes)
        out.append({"layer": k + 1, "z_top_mm": top, "z_bot_mm": bot,
                    "Psan": round(sand, 5), "Pcla": round(clay, 5),
                    "Porg": round(org, 5),
                    "source": hit["source"], "extrapolated": int(extrapolated),
                    "horizon_top_cm": hit["top_cm"], "horizon_bot_cm": hit["bot_cm"],
                    "Porg_reported": round(org_raw, 5),
                    "ksat_um_s": hit.get("ksat_um_s", ""),
                    "note": adjusted})
    return out, described_mm


def depth_weighted(layers: list[dict], key: str, to_mm: float | None = None) -> float:
    num = den = 0.0
    for l in layers:
        if to_mm is not None and l["z_top_mm"] >= to_mm:
            break
        bot = min(l["z_bot_mm"], to_mm) if to_mm is not None else l["z_bot_mm"]
        w = max(0.0, bot - l["z_top_mm"])
        num += w * l[key]
        den += w
    return num / den if den else float("nan")


# ------------------------------------------------------------------------ main

def process(st: dict, args, zr95: dict[str, float]) -> tuple[dict, list[dict], list[dict]]:
    sid, lat, lon = st["station_id"], st["lat"], st["lon"]
    info = {"station_id": sid, "lat": lat, "lon": lon,
            "forest_type": st["forest_type"], "note": ""}
    horizons: list[dict] = []
    ss: dict = {}

    order = [s.strip() for s in args.sources.split(",") if s.strip()]
    tried = []
    for src in order:
        if horizons:
            break
        tried.append(src)
        try:
            if src == "badm":
                horizons, f = badm_profile(args.badm_dir, sid)
                info["badm_file"] = f
            elif src == "ssurgo":
                ss = ssurgo_profile(lat, lon)
                horizons = ss.get("horizons") or []
            elif src == "polaris":
                horizons = polaris_profile(lat, lon)
            elif src == "soilgrids":
                horizons = soilgrids_profile(lat, lon)
        except Exception as exc:                                   # noqa: BLE001
            info["note"] = f"{src}: {type(exc).__name__}: {exc}"[:200]

    # The restriction is worth having even when the texture came from elsewhere,
    # since SSURGO is the only source that reports one.
    if not ss and "ssurgo" not in tried:
        try:
            ss = ssurgo_profile(lat, lon)
        except Exception:                                          # noqa: BLE001
            ss = {}

    info.update(texture_source=horizons[0]["source"] if horizons else "none",
                mukey=ss.get("mukey", ""), cokey=ss.get("cokey", ""),
                compname=ss.get("compname", ""), comppct=ss.get("comppct", ""),
                n_components=ss.get("n_components", ""),
                restriction=ss.get("restriction", ""),
                restriction_cm=ss.get("restriction_cm", ""),
                bedrock=ss.get("bedrock", ""),
                bedrock_cm=ss.get("bedrock_cm", ""),
                o_horizon_cm=ss.get("o_horizon_cm", ""),
                o_horizons=ss.get("o_horizons", ""),
                rootznaws_mm=ss.get("rootznaws_mm", ""),
                brockdepmin_cm=ss.get("brockdepmin_cm", ""))
    if ss.get("note"):
        info["note"] = ((info["note"] + "; ") if info["note"] else "") + ss["note"]
    if not horizons:
        info["status"] = "no_soil_data"
        return info, [], []

    described_cm = horizons[-1]["bot_cm"]
    res_cm = ss.get("restriction_cm")
    bed_cm = ss.get("bedrock_cm")
    zr = zr95.get(sid)

    # Column depth. corestrictions exists to record layers that restrict roots and
    # water -- NRCS already made that judgement in the field, which is the whole
    # reason for preferring SSURGO over a modelled bedrock raster. So ANY restriction
    # it reports sets the column, whatever its kind: there is no second opinion here
    # about whether paralithic bedrock or a fragipan "really" counts. The kind is
    # recorded so the choice stays visible, and --ignore-restrictions falls back to
    # the rooting-depth rule if a site ever needs it.
    #
    # min-depth is deliberately NOT applied to a restriction. Padding a 45 cm lithic
    # contact out to 1 m would invent water storage that is not there, and
    # plant-available water is the main soil control on the flux response this
    # project attributes to LMA.
    if bed_cm is not None and not args.ignore_bedrock:
        depth = bed_cm * 10.0
        rule = f"{ss['bedrock']} at {bed_cm:g} cm (SSURGO corestrictions)"
    else:
        depth = max(args.min_depth_mm, zr or 0.0, described_cm * 10.0)
        rule = (f"no bedrock reported; max(described {described_cm:g} cm, "
                f"ZR95 {zr or 0:g} mm, min-depth {args.min_depth_mm:g} mm)")
        if bed_cm is not None:
            rule += f"; {ss['bedrock']} at {bed_cm:g} cm ignored on request"
        if res_cm is not None:
            rule += f"; in-soil {ss['restriction']} at {res_cm:g} cm noted, not a floor"
    depth = min(depth, args.max_depth_mm)

    if zr and zr > depth:
        info["note"] = ((info["note"] + "; ") if info["note"] else "") + \
            f"ZR95 {zr:g} mm exceeds column {depth:g} mm -- T&C would abort; ZR95 must be capped"

    zs = build_zs(depth)
    # SSURGO leaves om_r null mostly in deep mineral C horizons, where organic matter
    # really is negligible. Filling those from the shallowest value -- which is the
    # organic-rich topsoil -- would hand a C horizon ~14% OM and, through Saxton &
    # Rawls, a wildly inflated porosity and water retention at depth. A small floor is
    # the far less wrong assumption, and n_org_filled makes it auditable.
    n_org_missing = sum(1 for h in horizons if h.get("org") is None)
    layers, described_mm = layerise(horizons, zs, ORG_FLOOR, args.max_porg)

    rows = [{"station_id": sid, **l} for l in layers]
    raw = [{"station_id": sid, **h} for h in horizons]
    n_extrap = sum(l["extrapolated"] for l in layers)
    info.update(status="ok", column_depth_mm=depth, depth_rule=rule,
                ms=len(layers), zs_mm=" ".join(f"{z:g}" for z in zs),
                described_to_cm=described_cm, n_horizons=len(horizons),
                n_horizons_no_org=n_org_missing,
                n_layers_porg_capped=sum(1 for l in layers
                                         if "capped" in (l.get("note") or "")),
                Porg_reported_max=round(max((h["org"] for h in horizons
                                             if h.get("org") is not None), default=0.0), 4),
                n_layers_extrapolated=n_extrap,
                ZR95_mm=zr if zr else "",
                Psan_mean=round(depth_weighted(layers, "Psan"), 4),
                Pcla_mean=round(depth_weighted(layers, "Pcla"), 4),
                Porg_mean=round(depth_weighted(layers, "Porg"), 4),
                Porg_top30=round(depth_weighted(layers, "Porg", 300.0), 4),
                Psan_top30=round(depth_weighted(layers, "Psan", 300.0), 4))
    return info, rows, raw


def probe_summary(rows: list[dict]) -> int:
    """Texture across the station set, from a --probe run.

    The point is to see the values before committing to a build: whether organic
    matter is actually reported, and whether texture varies enough with depth to
    justify layering rather than one triple per site.
    """
    have = [r for r in rows if r.get("n")]
    print(f"\n{'=' * 62}\nTEXTURE PROBE: {len(have)}/{len(rows)} stations with horizons"
          f"\n{'=' * 62}")
    if not have:
        return 1

    def stats(vals, name, fmt="%.3f"):
        v = sorted(x for x in vals if x is not None and math.isfinite(x))
        if not v:
            print(f"   {name:<22} no values")
            return
        print(f"   {name:<22} min {fmt % v[0]}  median {fmt % v[len(v) // 2]}  "
              f"max {fmt % v[-1]}")

    print("\nsurface horizon:")
    stats([r.get("top_sand") for r in have], "sand fraction")
    stats([r.get("top_clay") for r in have], "clay fraction")
    stats([r.get("top_org") for r in have], "Porg", "%.4f")
    print("\ndeepest horizon:")
    stats([r.get("deep_org") for r in have], "Porg", "%.4f")

    print("\nchange from surface to deepest horizon:")
    stats([r.get("sand_range") for r in have], "sand")
    stats([r.get("clay_range") for r in have], "clay")

    # If texture barely moves with depth, the layered profile buys little and the
    # simpler uniform column would have been defensible after all.
    moved = [r for r in have
             if abs(r.get("sand_range") or 0) > 0.05 or abs(r.get("clay_range") or 0) > 0.05]
    print(f"\nstations where sand or clay shifts by >0.05 with depth: "
          f"{len(moved)}/{len(have)}")

    noorg = sum(r.get("no_org", 0) for r in have)
    tot = sum(r["n"] for r in have)
    allmissing = sum(1 for r in have if r.get("no_org", 0) == r["n"])
    print(f"\nhorizons with no organic matter reported: {noorg}/{tot} "
          f"({100 * noorg / tot:.0f}%) -- filled with Porg = {ORG_FLOOR}")
    print(f"stations where NO horizon reports organic matter: {allmissing}")
    if allmissing:
        print("   " + " ".join(r["station_id"] for r in have
                               if r.get("no_org", 0) == r["n"])[:200])
    return 0


def report(out_dir: Path) -> int:
    """Summarise what was actually obtained for the station set."""
    sites_p, prof_p = out_dir / "soil_sites.csv", out_dir / "soil_profiles.csv"
    if not sites_p.is_file():
        print(f"no output to report on: {sites_p}", file=sys.stderr)
        return 1
    with open(sites_p, newline="", encoding="utf-8-sig") as fh:
        sites = list(csv.DictReader(fh))
    layers = []
    if prof_p.is_file():
        with open(prof_p, newline="", encoding="utf-8-sig") as fh:
            layers = list(csv.DictReader(fh))

    n = len(sites)
    ok = [s for s in sites if s.get("status") == "ok"]
    print(f"{'=' * 66}\nSOIL COVERAGE: {len(ok)}/{n} stations with a profile\n{'=' * 66}")

    print("\ntexture source (first hit in the chain):")
    for k, v in Counter(s.get("texture_source", "none") for s in sites).most_common():
        print(f"   {k:<18} {v:>4}  ({100 * v / n:.0f}%)")

    # Bedrock ends the soil column; everything else corestrictions reports sits
    # inside it. Reporting them together is what made this confusing to begin with.
    have_bed = [s for s in ok if s.get("bedrock_cm") not in MISSING]
    print(f"\nDEPTH TO BEDROCK (sets the column) : {len(have_bed)}/{len(ok)}")
    for k, v in Counter(s.get("bedrock", "") for s in have_bed).most_common():
        print(f"   {k or '(unnamed)':<28} {v:>4}")
    if have_bed:
        d = sorted(float(s["bedrock_cm"]) for s in have_bed)
        print(f"   depth cm: min {d[0]:.0f}  median {d[len(d) // 2]:.0f}  max {d[-1]:.0f}")
        print(f"   shallower than 100 cm: {sum(1 for x in d if x < 100)}")

    in_soil = [s for s in ok if s.get("restriction_cm") not in MISSING
               and s.get("bedrock_cm") in MISSING]
    print(f"\nin-soil restrictions (recorded, NOT a floor) : {len(in_soil)}")
    for k, v in Counter(s.get("restriction", "") for s in in_soil).most_common():
        print(f"   {k or '(unnamed)':<28} {v:>4}")
    print(f"\nno bedrock reported : {len(ok) - len(have_bed)}"
          f"  (column = max(described, ZR95, min-depth))")

    def nums(key, rows=ok):
        v = []
        for s in rows:
            try:
                v.append(float(s[key]))
            except (TypeError, ValueError, KeyError):
                pass
        return sorted(v)

    for key, unit in (("column_depth_mm", "mm"), ("described_to_cm", "cm"),
                      ("ms", "layers"), ("n_layers_extrapolated", "layers")):
        v = nums(key)
        if v:
            print(f"\n{key:<24} min {v[0]:g} {unit}   median {v[len(v) // 2]:g}   "
                  f"max {v[-1]:g}")

    deep = [s for s in ok if _f(s.get("column_depth_mm")) and float(s["column_depth_mm"]) > 2000]
    print(f"\ncolumns deeper than 2 m (beyond any source's description): {len(deep)}")
    shallow = [s for s in ok if _f(s.get("column_depth_mm")) and float(s["column_depth_mm"]) <= 1000]
    print(f"columns at or under 1 m                                  : {len(shallow)}")
    if shallow:
        print("   " + " ".join(s["station_id"] for s in shallow[:12]))

    tex = [(nums("Psan_mean"), "Psan"), (nums("Pcla_mean"), "Pcla"),
           (nums("Porg_top30"), "Porg(top 30cm)"), (nums("Porg_mean"), "Porg(column)")]
    print("\ntexture across stations:")
    for v, name in tex:
        if v:
            print(f"   {name:<16} min {v[0]:.3f}  median {v[len(v) // 2]:.3f}  max {v[-1]:.3f}")

    if layers:
        ex = sum(1 for l in layers if l.get("extrapolated") == "1")
        print(f"\nlayers written: {len(layers)}   extrapolated below the described "
              f"depth: {ex} ({100 * ex / len(layers):.0f}%)")
        print("layer source mix: " + ", ".join(
            f"{k} {v}" for k, v in Counter(l["source"] for l in layers).most_common()))

    notes = [s for s in sites if (s.get("note") or "").strip()]
    if notes:
        print(f"\n{len(notes)} station(s) with a note:")
        for s in notes[:15]:
            print(f"   {s['station_id']:<8} {s['note'][:96]}")
    bad = [s for s in sites if s.get("status") != "ok"]
    if bad:
        print(f"\n{len(bad)} station(s) WITHOUT a profile: "
              + " ".join(s["station_id"] for s in bad))
    return 0 if len(ok) == n else 1


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--site-list", type=Path, action="append", default=None)
    p.add_argument("--stations", default=None, help="comma-separated StationIDs")
    p.add_argument("--exclude-file", type=Path, default=DEFAULT_EXCLUDED)
    p.add_argument("--out", type=Path, default=DEFAULT_OUT)
    p.add_argument("--sources", default="badm,ssurgo,polaris,soilgrids",
                   help="source order, first hit wins (default: %(default)s)")
    p.add_argument("--badm-dir", type=Path, default=DEFAULT_BADM_DIR)
    p.add_argument("--root-depth", type=Path, default=DEFAULT_ROOT_DEPTH)
    p.add_argument("--ignore-bedrock", action="store_true",
                   help="do not let a reported bedrock contact set the column depth; "
                        "fall back to max(described, ZR95, --min-depth-mm) everywhere")
    p.add_argument("--max-porg", type=float, default=0.08,
                   help="Saxton & Rawls validity guard on Porg (default 0.08). O "
                        "horizons are removed by NAME, not by this threshold; the cap "
                        "only catches genuine mineral horizons above the ~8%% OM the "
                        "pedotransfer functions were fitted on, such as US-LPH's A "
                        "horizon at 10.8%%. Reported value kept in Porg_reported.")
    p.add_argument("--min-depth-mm", type=float, default=1000.0)
    p.add_argument("--max-depth-mm", type=float, default=5000.0)
    p.add_argument("--probe", action="store_true",
                   help="coverage only: query SDA per station, write nothing")
    p.add_argument("--report", action="store_true",
                   help="summarise an existing run in --out and exit")
    p.add_argument("--dry-run", action="store_true")
    args = p.parse_args()

    if args.report:
        return report(args.out)

    wanted = {s.strip() for s in args.stations.split(",")} if args.stations else None
    excluded = read_excluded(args.exclude_file)
    stations = [s for s in read_stations(args.site_list or DEFAULT_SITE_LISTS, wanted)
                if s["station_id"] not in excluded]
    if not stations:
        print("no stations to process", file=sys.stderr)
        return 1
    zr95 = read_root_depth(args.root_depth)

    print(f"stations   : {len(stations)} ({len(excluded)} excluded)")
    print(f"sources    : {args.sources}")
    print(f"ZR95 known : {len(zr95)} station(s) from {args.root_depth}")
    print(f"depth      : {args.min_depth_mm:g}-{args.max_depth_mm:g} mm")
    print(f"output     : {args.out}\n")
    if args.dry_run:
        for s in stations:
            print(f"  - {s['station_id']:<8} {s['lat']:9.4f} {s['lon']:10.4f}")
        return 0

    sites, profiles, raws, probe_rows = [], [], [], []
    for i, st in enumerate(stations, 1):
        if args.probe:
            try:
                ss = ssurgo_profile(st["lat"], st["lon"])
                hz = ss.get("horizons") or []
                # Name the component when there are no horizons: a map unit can be a
                # miscellaneous area (Water, Rock outcrop, Urban land) or an
                # unpopulated NOTCOM survey, which is a real SSURGO gap and a correct
                # fall-through to POLARIS -- not the same thing as a query bug.
                if not hz:
                    print(f"  {st['station_id']:<8} mukey {str(ss.get('mukey', '-')):<10} "
                          f" 0 horizons   <-- no horizons: component "
                          f"'{ss.get('compname') or '?'}'")
                    probe_rows.append({"station_id": st["station_id"], "n": 0})
                    continue
                top, bot = hz[0], hz[-1]
                n_noorg = sum(1 for h in hz if h.get("org") is None)
                print(f"  {st['station_id']:<8} mukey {str(ss.get('mukey', '-')):<10} "
                      f"{len(hz):>2} hz to {bot['bot_cm']:>4.0f} cm  "
                      f"top san/cla/org {top['sand']:.2f}/{top['clay']:.2f}/"
                      f"{(top['org'] if top['org'] is not None else float('nan')):.4f}  "
                      f"deep {bot['sand']:.2f}/{bot['clay']:.2f}/"
                      f"{(bot['org'] if bot['org'] is not None else float('nan')):.4f}  "
                      f"no-org {n_noorg}/{len(hz)}  "
                      f"{ss.get('bedrock') or ss.get('restriction') or '-'}")
                probe_rows.append({
                    "station_id": st["station_id"], "n": len(hz), "no_org": n_noorg,
                    "top_sand": top["sand"], "top_clay": top["clay"],
                    "top_org": top["org"], "deep_org": bot["org"],
                    "sand_range": bot["sand"] - top["sand"],
                    "clay_range": bot["clay"] - top["clay"]})
            except Exception as exc:                               # noqa: BLE001
                print(f"  {st['station_id']:<8} ! {type(exc).__name__}: {exc}")
            continue
        info, rows, raw = process(st, args, zr95)
        sites.append(info)
        profiles.extend(rows)
        raws.extend(raw)
        print(f"  [{i:>3}/{len(stations)}] {info['station_id']:<8} "
              f"{info.get('texture_source', '-'):<14} "
              f"depth {info.get('column_depth_mm', '-'):>6} mm  "
              f"ms {info.get('ms', '-'):>2}  "
              f"extrap {info.get('n_layers_extrapolated', '-'):>2}  "
              f"{info.get('restriction', '') or ''}")
    if args.probe:
        return probe_summary(probe_rows)

    args.out.mkdir(parents=True, exist_ok=True)
    for name, rows in (("soil_sites.csv", sites), ("soil_profiles.csv", profiles),
                       ("soil_horizons_raw.csv", raws)):
        if not rows:
            continue
        keys = list({k: None for r in rows for k in r})
        with open(args.out / name, "w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=keys, extrasaction="ignore")
            w.writeheader()
            w.writerows(rows)
        print(f"wrote {args.out / name}  ({len(rows)} rows)")

    with open(args.out / "soil_provenance.json", "w", encoding="utf-8") as fh:
        json.dump({
            "source_order": args.sources,
            "sources": {
                "ameriflux_badm": "in-situ SOIL_TEX groups from the station BIF file",
                "ssurgo_sda": f"USDA Soil Data Access ({SDA_URL}), dominant component "
                              f"by comppct_r; horizons from chorizon, restriction from "
                              f"corestrictions then muaggatt.brockdepmin",
                "polaris": f"{POLARIS_BASE} 30 m CONUS, mean, 6 standard intervals",
                "soilgrids": f"{SOILGRIDS_URL} 250 m global, mean",
            },
            "organic_matter": "Porg is a FRACTION. SSURGO om_r is percent; POLARIS om "
                              "is log10(percent); SoilGrids soc is dg/kg and is "
                              "converted with OM = SOC * 1.72.",
            "layering": "texture mapped onto Zs by layer midpoint; Soil_parameters is "
                        "run per layer. No source resolves finer than 5 cm, so the top "
                        "three layers share a value.",
            "extrapolation": "below the described depth the deepest observed layer is "
                             "carried down; those layers are flagged `extrapolated`.",
            "depth_rule":
                "Depth to BEDROCK from SSURGO corestrictions (lithic/paralithic/densic "
                "bedrock) for the dominant component sets the column. Other restriction "
                "kinds -- densic material, fragipans, abrupt textural changes -- lie "
                "WITHIN the soil and are recorded but never used as a floor: the survey "
                "keeps describing horizons 80-140 cm below them, whereas at a bedrock "
                "contact the description stops dead. Where no bedrock is reported, "
                f"max(described depth, ZR95, {args.min_depth_mm:g} mm), capped at "
                f"{args.max_depth_mm:g} mm. Rooting depth comes from Schenk & Jackson, "
                "not from here; this step supplies the soil column only. "
                "muaggatt.brockdepmin is recorded but never used -- it is the minimum "
                "over all components of the map unit and contradicted the dominant "
                "component at every station where it fired. rootznaws is recorded as a "
                "water-storage cross-check; rootznemc is deliberately not used, being a "
                "capped commodity-crop rooting depth that under-reports forests.",
            "sand_clay_convention": "sandtotal_r/claytotal_r are percentages of the "
                                    "<2 mm mineral fraction; converted to fractions and "
                                    "scaled down only if Psan+Pcla+Porg would exceed 1, "
                                    "which Soil_parameters rejects.",
        }, fh, indent=2)

    print()
    return report(args.out)


if __name__ == "__main__":
    sys.exit(main())
