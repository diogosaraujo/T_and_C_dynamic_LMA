#!/usr/bin/env python3
"""Report which T&C parameters the downloaded AmeriFlux BADM can actually supply.

BADM coverage is uneven: most sites report a few groups well and leave the rest empty,
and the naming is not identical across sites. So this DISCOVERS what each station reports
rather than assuming a schema -- it matches variables against the wish list in
ameriflux_api.PARAMETER_TARGETS and separately inventories everything it did not match.
An unexpected or renamed variable therefore shows up as "unmatched", not as "absent".

Outputs, next to the downloaded data:

    badm_coverage.csv     station x parameter -> yes/no, with the value(s) found
    badm_values.csv       every matched variable and its value (long format)
    badm_inventory.csv    every variable group/variable present, matched or not

and a console summary of how many stations can supply each parameter.

    python inspect_ameriflux_badm.py
    python inspect_ameriflux_badm.py --stations US-HBK,US-Ha2
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from collections import defaultdict
from pathlib import Path

from ameriflux_api import PARAMETER_TARGETS

INPUT_ROOT = Path(os.environ.get("TC_INPUT_DATA", "/vol_efthymios/NFS07/dd1136/T_and_C/input_data"))
DEFAULT_DIR = INPUT_ROOT / "ameriflux"

# BIF long format: one row per reported variable.
REQUIRED_COLS = ("SITE_ID", "VARIABLE_GROUP", "VARIABLE", "DATAVALUE")
MISSING_VALUES = {"", "-9999", "-9999.0", "NA", "NaN", "None"}


def log(msg: str = "") -> None:
    print(msg, flush=True)


def read_badm_rows(path: Path) -> list[dict]:
    """Read a BADM/BIF table from .xlsx or .csv into dicts keyed by column name."""
    if path.suffix.lower() in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise SystemExit(
                "openpyxl is required to read BADM .xlsx files. "
                "Run:  pip install -r requirements.txt"
            )
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            rows_iter = ws.iter_rows(values_only=True)
            header = None
            out = []
            for raw in rows_iter:
                values = ["" if v is None else str(v).strip() for v in raw]
                if header is None:
                    # Tolerate preamble rows above the real header.
                    if any(c.upper() == "SITE_ID" for c in values):
                        header = [c.upper() for c in values]
                    continue
                out.append(dict(zip(header, values)))
            return out
        finally:
            wb.close()

    with path.open(newline="", encoding="utf-8-sig", errors="replace") as fh:
        reader = csv.DictReader(fh)
        return [{(k or "").upper(): (v or "").strip() for k, v in row.items()} for row in reader]


def find_badm_files(station_dir: Path) -> list[Path]:
    """Locate BADM/BIF tables, without mistaking a BASE measurement file for one.

    The product token sits between underscores in AmeriFlux filenames
    (AMF_<SITE>_BIF_<POLICY>_<VERSION>.xlsx), so match on the delimited token rather
    than a bare substring -- otherwise a site whose ID contains 'BIF'/'BADM' would have
    its BASE csv picked up as metadata.
    """
    hits = []
    for path in sorted(station_dir.iterdir()):
        if not path.is_file() or path.suffix.lower() not in (".xlsx", ".xls", ".csv"):
            continue
        tokens = {t.upper() for t in path.stem.split("_")}
        if tokens & {"BIF", "BADM"} and "BASE" not in tokens:
            hits.append(path)
    return hits


def match_targets(rows: list[dict]) -> tuple[dict[str, list[dict]], list[dict]]:
    """Split reported variables into wish-list matches and everything else."""
    matched: dict[str, list[dict]] = defaultdict(list)
    unmatched: list[dict] = []

    for row in rows:
        variable = (row.get("VARIABLE") or "").strip()
        value = (row.get("DATAVALUE") or "").strip()
        if not variable or value in MISSING_VALUES:
            continue
        group = (row.get("VARIABLE_GROUP") or "").strip()
        hay = f"{group} {variable}".upper()

        hit = None
        for target in PARAMETER_TARGETS:
            if any(pat.upper() in hay for pat in target["patterns"]):
                hit = target["key"]
                break
        entry = {"variable_group": group, "variable": variable, "value": value}
        if hit:
            matched[hit].append(entry)
        else:
            unmatched.append(entry)
    return matched, unmatched


def summarise(values: list[dict], limit: int = 3) -> str:
    """A short, readable digest of the values found for one parameter."""
    uniq: list[str] = []
    for entry in values:
        v = entry["value"]
        if v not in uniq:
            uniq.append(v)
        if len(uniq) >= limit:
            break
    text = "; ".join(uniq)
    extra = len({e["value"] for e in values}) - len(uniq)
    return text + (f" (+{extra} more)" if extra > 0 else "")


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(
        description="Report which T&C parameters AmeriFlux BADM can supply.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--dir", type=Path, default=DEFAULT_DIR, help="ameriflux output directory")
    p.add_argument("--stations", default=None, help="comma-separated StationIDs")
    p.add_argument("--show-unmatched", type=int, default=12,
                   help="how many unmatched variable groups to list per station (0 = none)")
    args = p.parse_args(argv)

    if not args.dir.is_dir():
        raise SystemExit(f"not a directory: {args.dir}  (run download_ameriflux.py first)")

    wanted = {s.strip() for s in args.stations.split(",") if s.strip()} if args.stations else None
    station_dirs = sorted(d for d in args.dir.iterdir()
                          if d.is_dir() and not d.name.startswith("_"))
    if wanted is not None:
        station_dirs = [d for d in station_dirs if d.name in wanted]
    if not station_dirs:
        raise SystemExit(f"no station directories in {args.dir}")

    # Site-level metadata fills in elevation/IGBP even when BADM does not report them.
    site_meta = {}
    meta_path = args.dir / "site_metadata.json"
    if meta_path.exists():
        site_meta = json.loads(meta_path.read_text(encoding="utf-8"))

    coverage_rows, value_rows, inventory_rows = [], [], []
    per_target_count: dict[str, int] = defaultdict(int)
    n_with_badm = 0

    log(f"inspecting {len(station_dirs)} station(s) in {args.dir}\n")

    for station_dir in station_dirs:
        sid = station_dir.name
        badm_files = find_badm_files(station_dir)
        if not badm_files:
            log(f"  {sid}: no BADM file found")
            coverage_rows.append({"station_id": sid, "has_badm": "no",
                                  **{t["key"]: "" for t in PARAMETER_TARGETS}})
            continue

        rows: list[dict] = []
        for path in badm_files:
            try:
                rows.extend(read_badm_rows(path))
            except Exception as exc:
                log(f"  ! {sid}: cannot read {path.name} ({exc})")
        if not rows:
            log(f"  {sid}: BADM file present but no readable rows")
            continue
        missing_cols = [c for c in REQUIRED_COLS if c not in rows[0]]
        if missing_cols:
            log(f"  ! {sid}: BADM table missing expected column(s) {missing_cols}; "
                f"found {sorted(rows[0])[:8]}")

        n_with_badm += 1
        matched, unmatched = match_targets(rows)

        cov = {"station_id": sid, "has_badm": "yes"}
        for target in PARAMETER_TARGETS:
            key = target["key"]
            hits = matched.get(key, [])
            cov[key] = summarise(hits) if hits else ""
            if hits:
                per_target_count[key] += 1
            for entry in hits:
                value_rows.append({"station_id": sid, "parameter": key, **entry})

        # Elevation and IGBP usually come from the site registry, not BADM.
        rec = site_meta.get(sid, {})
        if not cov.get("elevation"):
            elev = (rec.get("GRP_LOCATION") or {}).get("LOCATION_ELEV")
            if elev not in (None, ""):
                cov["elevation"] = f"{elev} (site registry)"
                per_target_count["elevation"] += 1
        if not cov.get("igbp") and rec.get("IGBP"):
            cov["igbp"] = f"{rec['IGBP']} (site registry)"
            per_target_count["igbp"] += 1

        coverage_rows.append(cov)

        groups = sorted({e["variable_group"] for e in unmatched})
        for entry in unmatched:
            inventory_rows.append({"station_id": sid, "matched": "no", **entry})
        for key, hits in matched.items():
            for entry in hits:
                inventory_rows.append({"station_id": sid, "matched": key, **entry})

        have = [t["key"] for t in PARAMETER_TARGETS if cov.get(t["key"])]
        log(f"  {sid}: {len(rows)} BADM rows | supplies: {', '.join(have) if have else 'nothing on the wish list'}")
        if args.show_unmatched and groups:
            shown = groups[: args.show_unmatched]
            more = len(groups) - len(shown)
            log(f"      other groups reported: {', '.join(shown)}"
                + (f" (+{more} more)" if more > 0 else ""))

    # ---------------- outputs ----------------
    if coverage_rows:
        fields = ["station_id", "has_badm"] + [t["key"] for t in PARAMETER_TARGETS]
        with (args.dir / "badm_coverage.csv").open("w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
            w.writeheader()
            w.writerows(coverage_rows)
    for name, rows_out, fields in (
        ("badm_values.csv", value_rows,
         ["station_id", "parameter", "variable_group", "variable", "value"]),
        ("badm_inventory.csv", inventory_rows,
         ["station_id", "matched", "variable_group", "variable", "value"]),
    ):
        if rows_out:
            with (args.dir / name).open("w", newline="", encoding="utf-8") as fh:
                w = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
                w.writeheader()
                w.writerows(rows_out)

    def fit(text: str, width: int) -> str:
        return text if len(text) <= width else text[: width - 1] + "…"

    total = len(station_dirs)
    log(f"\n{'parameter':<16} {'sites':>8}   {'T&C use':<52}  fallback when absent")
    log("-" * 140)
    for target in PARAMETER_TARGETS:
        n = per_target_count.get(target["key"], 0)
        pct = f"{n}/{total}"
        log(f"{target['key']:<16} {pct:>8}   {fit(target['tc_use'], 52):<52}  "
            f"{fit(target['fallback'], 58)}")

    log(f"\n{n_with_badm}/{total} station(s) had a readable BADM file")
    log(f"written: badm_coverage.csv, badm_values.csv, badm_inventory.csv (in {args.dir})")
    log("\nbadm_inventory.csv lists every variable reported, including ones outside the")
    log("wish list -- check it before concluding a parameter is unavailable.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
