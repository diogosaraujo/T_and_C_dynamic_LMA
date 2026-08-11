#!/usr/bin/env python3
"""Compare the PLSR LMA driving the model against measured LMA in AmeriFlux BADM.

LMA is the treatment variable: it enters T&C only as Sl = 1/(LMA*f_C) and drives
LAI = Sl*B(1), so a bias here is worse than a bias in any parameter. At US-Ha2 the
PLSR series (mean 102.6, range 78-119 g/m2) sits 22% BELOW the BADM measurement of
131 g/m2 -- and 131 lies outside the entire 37-year PLSR distribution, so it is an
offset rather than a missed year. Substituting the measured value reproduces the
reported LAI of 4.4 almost exactly, which means the model's carbon allocation is
right and the input is not.

This asks how general that is:
  * which stations report LMA in BADM at all
  * how the measurement compares with the PLSR mean and with the full series
  * whether any offset is systematic, evergreen-specific, or scattered

Scattered differences would point at spatial mismatch (a ~9 km ERA5-Land pixel
against a tower footprint) and belong in the uncertainty discussion; a consistent
offset would be a correction to apply to the LMA series.

    python check_lma_vs_badm.py                 # summary
    python check_lma_vs_badm.py --csv out.csv   # plus a per-station dump

Reads only what is already on disk: the BADM files download_ameriflux.py fetched
and the LMA_<ST>.mat files build_model_run.py generated.
"""
from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from pathlib import Path

import numpy as np

REPO_ROOT = Path(__file__).resolve().parents[1]
INPUT_ROOT = Path(os.environ.get("TC_INPUT_DATA",
                                 "/vol_efthymios/NFS07/dd1136/T_and_C/input_data"))
DEFAULT_BADM = INPUT_ROOT / "ameriflux"
DEFAULT_MODEL_RUN = Path(os.environ.get("MODEL_RUN", INPUT_ROOT.parent / "model_run"))
DEFAULT_SITE_LISTS = [
    REPO_ROOT / "T&C" / "dynamic_lma_test" / "deciduous_ameriflux.csv",
    REPO_ROOT / "T&C" / "dynamic_lma_test" / "evergreen_ameriflux.csv",
]
F_C = 0.5          # LMA is DRY MASS (confirmed 2026-08-11); 0.5 converts to gC

# BADM reports the value as LMA and qualifies it with LMA_COMMENT / LMA_SPP / units.
# Match the value itself, not its qualifiers -- an unanchored "LMA" also catches
# LMA_COMMENT and would turn a text note into a number.
LMA_VALUE = re.compile(r"^LMA$", re.I)
LMA_QUAL = re.compile(r"^LMA_", re.I)


def read_site_list(paths) -> dict[str, str]:
    out = {}
    for p in paths:
        if not Path(p).is_file():
            print(f"  ! site list not found: {p}", file=sys.stderr)
            continue
        for r in csv.DictReader(open(p, newline="", encoding="utf-8-sig")):
            sid = (r.get("StationID") or r.get("SITE_ID") or "").strip()
            if sid:
                out[sid] = (r.get("ForestType") or "").strip().lower()
    return out


def badm_lma(station_dir: Path):
    """Measured LMA and its qualifiers, or None. Returns (value, {qualifiers})."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("openpyxl is required; it is already in requirements.txt")
    files = [p for p in sorted(station_dir.iterdir())
             if p.is_file() and p.suffix.lower() in (".xlsx", ".xls")
             and {t.upper() for t in p.stem.split("_")} & {"BIF", "BADM"}]
    for path in files:
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception:                                   # noqa: BLE001
            continue
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        try:
            next(rows)
        except StopIteration:
            continue
        val, qual = None, {}
        for r in rows:
            if not r or len(r) < 5:
                continue
            var, dv = str(r[3] or "").strip(), r[4]
            if LMA_VALUE.match(var):
                try:
                    val = float(str(dv).strip())
                except (TypeError, ValueError):
                    pass
            elif LMA_QUAL.match(var):
                qual[var.upper()] = str(dv).strip()
        if val is not None:
            return val, qual
    return None, {}


def plsr_series(model_run: Path, station: str):
    """The PLSR LMA series the model was driven with, or None."""
    m = model_run / station / "era5_land" / "dyn_lma" / f"LMA_{station.replace('-', '_')}.mat"
    if not m.is_file():
        return None, None
    try:
        from scipy.io import loadmat
        d = loadmat(m)
        lma = np.asarray(d["LMA"], dtype=float).ravel()
        yrs = np.asarray(d["years"]).ravel().astype(int)
    except Exception:                                       # noqa: BLE001
        try:
            import h5py
            with h5py.File(m, "r") as f:
                lma = np.array(f["LMA"]).ravel().astype(float)
                yrs = np.array(f["years"]).ravel().astype(int)
        except Exception:                                   # noqa: BLE001
            return None, None
    good = np.isfinite(lma)
    return lma[good], yrs[good] if len(yrs) == len(good) else None


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--badm", type=Path, default=DEFAULT_BADM)
    ap.add_argument("--model-run", type=Path, default=DEFAULT_MODEL_RUN)
    ap.add_argument("--site-list", type=Path, nargs="*", default=None)
    ap.add_argument("--csv", type=Path, default=None)
    a = ap.parse_args(argv)

    if not a.badm.is_dir():
        print(f"ERROR: BADM directory not found: {a.badm}", file=sys.stderr)
        return 1
    wanted = read_site_list(a.site_list or DEFAULT_SITE_LISTS)
    print(f"BADM      : {a.badm}")
    print(f"model_run : {a.model_run}")
    print(f"site list : {len(wanted)} stations\n")

    rows, n_dirs = [], 0
    for d in sorted(p for p in a.badm.iterdir() if p.is_dir()):
        sid = d.name
        if wanted and sid not in wanted:
            continue
        n_dirs += 1
        meas, qual = badm_lma(d)
        series, _ = plsr_series(a.model_run, sid)
        rows.append({"StationID": sid, "ForestType": wanted.get(sid, ""),
                     "badm_LMA": meas, "series": series, "qual": qual})

    have_m = [r for r in rows if r["badm_LMA"] is not None]
    have_b = [r for r in have_m if r["series"] is not None and len(r["series"])]
    print(f"{'=' * 78}\n{n_dirs} stations scanned\n"
          f"  {len(have_m)} report LMA in BADM\n"
          f"  {len(have_b)} have BOTH a measurement and a PLSR series\n{'=' * 78}")
    if not have_b:
        print("\nNothing comparable. If stations report LMA but have no series, the model_run"
              "\ntree has not been built for them yet.")
        return 0

    print(f"\n  {'station':<9}{'type':<10}{'BADM':>7}{'PLSR mean':>10}{'diff %':>8}"
          f"{'PLSR range':>14}{'pctile':>8}")
    for r in sorted(have_b, key=lambda x: x["StationID"]):
        s = r["series"]; m = r["badm_LMA"]
        pct = 100.0 * float(np.mean(s < m))
        print(f"  {r['StationID']:<9}{r['ForestType'][:9]:<10}{m:>7.1f}{s.mean():>10.1f}"
              f"{100*(s.mean()/m - 1):>+7.0f}%{f'{s.min():.0f}-{s.max():.0f}':>14}"
              f"{pct:>7.0f}%")
    print("\n  diff % = PLSR mean relative to the measurement (negative = PLSR too low)")
    print("  pctile = where the measurement falls in the PLSR distribution;")
    print("           0 or 100 means it lies outside the simulated range entirely")

    print(f"\n{'=' * 78}\nIS THE OFFSET SYSTEMATIC?\n{'=' * 78}")
    for ft in sorted({r["ForestType"] for r in have_b}):
        g = [r for r in have_b if r["ForestType"] == ft]
        d = np.array([100*(r["series"].mean()/r["badm_LMA"] - 1) for r in g])
        out = sum(1 for r in g
                  if r["badm_LMA"] < r["series"].min() or r["badm_LMA"] > r["series"].max())
        print(f"  {ft or '?':<10} n={len(g):<3} mean {d.mean():+6.1f}%  median {np.median(d):+6.1f}%"
              f"  sd {d.std():5.1f}  range {d.min():+.0f}..{d.max():+.0f}%")
        print(f"  {'':<10} {out}/{len(g)} measurements fall OUTSIDE the PLSR series range")
        if len(g) >= 3:
            same = (d < 0).all() or (d > 0).all()
            print(f"  {'':<10} -> {'CONSISTENT sign: looks like a retrieval offset' if same else 'MIXED signs: looks like spatial mismatch, not a common bias'}")

    print("\n  Interpreting this:")
    print("   * consistent sign + tight sd  -> a correction to apply to the LMA series")
    print("   * mixed signs / wide sd       -> ~9 km pixel vs tower footprint; report as")
    print("                                    uncertainty rather than correcting")
    print("   * an offset in EVERGREEN only -> a needleleaf retrieval problem, which would")
    print("                                    affect 53 of the 98 stations")

    if a.csv:
        a.csv.parent.mkdir(parents=True, exist_ok=True)
        with open(a.csv, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["StationID", "ForestType", "badm_LMA_g_m2", "plsr_mean", "plsr_min",
                        "plsr_max", "plsr_n_years", "diff_pct", "measured_pctile_of_series",
                        "Sl_from_badm", "Sl_from_plsr_mean", "LMA_COMMENT", "LMA_SPP"])
            for r in sorted(rows, key=lambda x: x["StationID"]):
                s, m = r["series"], r["badm_LMA"]
                w.writerow([
                    r["StationID"], r["ForestType"], "" if m is None else f"{m:.2f}",
                    "" if s is None or not len(s) else f"{s.mean():.2f}",
                    "" if s is None or not len(s) else f"{s.min():.2f}",
                    "" if s is None or not len(s) else f"{s.max():.2f}",
                    "" if s is None else len(s),
                    "" if (m is None or s is None or not len(s)) else f"{100*(s.mean()/m-1):.1f}",
                    "" if (m is None or s is None or not len(s)) else f"{100*np.mean(s<m):.0f}",
                    "" if m is None else f"{1/(m*F_C):.5f}",
                    "" if s is None or not len(s) else f"{1/(s.mean()*F_C):.5f}",
                    r["qual"].get("LMA_COMMENT", "").replace("\n", " ")[:200],
                    r["qual"].get("LMA_SPP", "")])
        print(f"\nper-station dump -> {a.csv}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
